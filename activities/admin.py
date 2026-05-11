from io import BytesIO
from django.contrib import admin
from django.contrib.auth.models import Group
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .models import Submission

# Remove Groups from admin sidebar
admin.site.unregister(Group)


@admin.register(Submission)
class SubmissionAdmin(admin.ModelAdmin):
    """Clean admin panel for managing student submissions."""
    list_display = ('full_name', 'student_id', 'section', 'choice', 'created_at')
    list_filter = ('choice', 'section')
    search_fields = ('full_name', 'student_id')
    actions = ['export_as_excel']
    list_per_page = 50
    ordering = ('-created_at',)

    def export_as_excel(self, request, queryset):
        """Export selected submissions to Excel (.xlsx)."""
        wb = Workbook()
        ws = wb.active
        ws.title = 'Submissions'

        # Header styling
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        headers = ['Full Name', 'Student ID', 'Section', 'Activity', 'Viva Topic', 'Submitted At']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Data rows
        for row_num, obj in enumerate(queryset, 2):
            row_data = [
                obj.full_name,
                obj.student_id,
                obj.section,
                obj.get_choice_display(),
                obj.viva_topic or '-',
                obj.created_at.strftime('%Y-%m-%d %H:%M'),
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

        # Auto-fit column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 4

        # Save to response
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="submissions.xlsx"'
        return response

    export_as_excel.short_description = 'Export selected to Excel'
